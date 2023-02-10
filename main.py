import cfscrape
import xlsxwriter
from fake_useragent import UserAgent
from bs4 import BeautifulSoup
from openpyxl.reader.excel import load_workbook

user = UserAgent().random

header = {
    'user-agent': user
}


def get_data(urs):
    try:
        page_count = 1
        with cfscrape.Session() as session:
            for u in urs:

                print('Page ' + str(page_count))
                link = u
                check_response = session.get(url=link, headers=header)

                with open('check.html', 'w', encoding='utf-8') as file:
                    file.write(check_response.text)

                print(check_response.status_code, check_response.reason)
                page_count += 1
                yield check_response

    except Exception as ex:
        print(ex)


def pars_data(urs: tuple):
    data_some = []

    try:
        for data in get_data(urs):
            bs = BeautifulSoup(data.text, 'lxml')

            col = bs.find_all('tr')

            len_col = len(col)

            for l in range(len_col):

                card = col[l].text.replace('\n', ' ')

                if 'Марка' in card:
                    name = bs.find('h1', class_='h1').text
                    data_some.append(name)

                    sort_brand = card.split('Марка')[1].replace(' ', '')
                    data_some.append(sort_brand)

                if 'Модель' in card:
                    sort_model = card.split('Модель')[1].replace('на сайте вендора Перейти на', '').replace('сайт >>>', '').replace(' ', '')
                    data_some.append(sort_model)

                if 'Страна производства' in card:
                    sort_coutry = card.split('Страна производства')[1].replace(' ', '')
                    data_some.append(sort_coutry)

                if 'Габариты упаковки' in card:
                    sort_dimension_box = card.split('ШхВхГ')[1].replace(')', '').replace(' ', '')
                    data_some.append(sort_dimension_box)

                if 'Габариты прибора' in card:
                    sort_dimension_instrument = card.split('ШхВхГ')[1].replace(')', '').replace(' ', '')
                    data_some.append(sort_dimension_instrument)

                if 'Масса, брутто' in card:
                    weight_instrument = card.split('(кг)')[1].replace(' ', '')
                    data_some.append(weight_instrument)

                if 'Масса, нетто' in card:
                    weight_box = card.split('(кг)')[1].replace(' ', '')
                    data_some.append(weight_box)

                if 'EAN' in card:
                    ean = card.split('EAN')[1].replace(' ', '')
                    data_some.append(ean)

                if 'Артикул вендора' in card:
                    art_vendor = card.split('Артикул вендора')[1].replace(' ', '')
                    data_some.append(art_vendor)

                if 'Артикул 1С' in card:
                    art_1c = card.split('Артикул 1С')[1].replace(' ', '')
                    data_some.append(art_1c)

            yield data_some

            data_some.clear()

    except Exception as ex:
        print(ex)


def save_data(urs: tuple):
    try:
        file_name = 'content/content_2.xlsx'

        book = xlsxwriter.Workbook(file_name)  # Создаем файл Exel
        page = book.add_worksheet('content')

        row = 0

        page.set_column('A:A', 25)
        page.set_column('B:B', 25)
        page.set_column('C:C', 25)
        page.set_column('D:D', 25)
        page.set_column('E:E', 25)
        page.set_column('F:F', 20)
        page.set_column('G:G', 20)
        page.set_column('H:H', 15)
        page.set_column('I:I', 20)
        page.set_column('J:J', 15)
        page.set_column('K:K', 15)

        page.write(0, 0, 'Наименование')
        page.write(0, 0 + 1, 'Марка')
        page.write(0, 0 + 2, 'Модель')
        page.write(0, 0 + 3, 'Страна производства')
        page.write(0, 0 + 4, 'Габариты прибора(ШхВхГ)')
        page.write(0, 0 + 5, 'Габариты упаковки(ШхВхГ)')
        page.write(0, 0 + 6, 'Масса, брутто (кг)')
        page.write(0, 0 + 7, 'Масса, нетто (кг)')
        page.write(0, 0 + 8, 'EAN')
        page.write(0, 0 + 9, 'Артикул вендора')
        page.write(0, 0 + 10,  'Артикул 1С')

        row += 1

        book.close()

        xl_file = load_workbook(file_name)
        page = xl_file['content']

        try:
            for data in pars_data(urs):
                try:
                    page.append([data[0], data[1], data[2], data[3], data[4], data[5], data[6], data[7], data[8], data[9], data[10]])
                except Exception as ex:
                    print(ex)
                    page.append(['Ошибка'])


        except Exception as ex:
            print(ex)
            page.append(['Ошибка'])


        xl_file.save(file_name)
        xl_file.close()

    except Exception as ex:
        print(ex)
